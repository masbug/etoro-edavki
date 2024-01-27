#!/usr/bin/python

# python 3.10+ fix
import sys
if sys.version_info.major == 3 and sys.version_info.minor >= 10:
    import collections.abc
    collections.Iterable = collections.abc.Iterable

import urllib.request
import sys
import xml.etree.ElementTree
import datetime
import os
import glob
import argparse
#import locale
#import prettytable
from xml.dom import minidom

from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl_templates import TemplatedWorkbook
from openpyxl_templates.styles import DefaultStyleSet
from openpyxl_templates.table_sheet import TableSheet
from openpyxl_templates.table_sheet.columns import CharColumn
from operator import itemgetter

APP_VER = "1.6.4"

EDAVKI_DATETIME_FORMAT = "%Y-%m-%d"
ETORO_DATETIME_FORMAT_EN1 = "%d/%m/%Y %H:%M:%S"
ETORO_DATETIME_FORMAT_EN2 = "%d/%m/%Y"
ETORO_DATETIME_FORMAT_SL1 = "%d.%m.%Y %H:%M:%S"
ETORO_DATETIME_FORMAT_SL2 = "%d.%m.%Y"
ETORO_CURRENCY = "USD"

bsRateXmlUrl = "https://www.bsi.si/_data/tecajnice/dtecbs-l.xml"
ignoreAssets = []
derivateAssets = ["CFD", "OPT", "FUT", "FOP"]
normalAssets = ["Stocks", "Crypto", "ETF"]

dividendMarker = "Payment caused by dividend"

float_with_comma = False

class ClosedPositionsSheet(TableSheet):
    # 2024: Position ID	Action	Amount	Units	Open Date	Close Date	Leverage	Spread Fees (USD)	Profit(USD)	Profit(EUR)	Open Rate
    #       Close Rate	Take profit rate	Stop lose rate	Rollover Fees and Dividends	Copied From	Type	ISIN	Notes
    position_id = CharColumn(header="Position ID")
    action = CharColumn(header="Action")
    amount = CharColumn(header="Amount")
    units = CharColumn(header="Units")
    open_date = CharColumn(header="Open Date")
    close_date = CharColumn(header="Close Date")
    leverage = CharColumn(header="Leverage")
    spread = CharColumn(header="Spread Fees (USD)")
    market_spread = CharColumn(header="Market Spread (USD)")
    profit = CharColumn(header="Profit(USD)")
    profit_eur = CharColumn(header="Profit(EUR)")
    open_rate = CharColumn(header="Open Rate")
    close_rate = CharColumn(header="Close Rate")
    take_profit_rate = CharColumn(header="Take profit rate")
    stop_loss_rate = CharColumn(header="Stop lose rate")
    rollover_fees_and_dividends = CharColumn(header="Rollover Fees and Dividends")
    trader = CharColumn(header="Copied From")
    type = CharColumn(header="Type")
    isin = CharColumn(header="ISIN")
    notes = CharColumn(header="Notes")

class AccountActivityReportSheet(TableSheet):
    # 2022: Date	Type	Details	Amount	Realized Equity Change	Realized Equity	Balance	Position ID	NWA
    # 2023: Date	Type	Details	Amount	Units	Realized Equity Change	Realized Equity	Balance	Position ID	Asset type	NWA
    date = CharColumn(header="Date")
    type = CharColumn(header="Type")
    details = CharColumn(header="Details")
    amount = CharColumn(header="Amount")
    units = CharColumn(header="Units")
    realized_equity_change = CharColumn(header="Realized Equity Change")
    realized_equity = CharColumn(header="Realized Equity")
    account_balance = CharColumn(header="Balance")
    position_id = CharColumn(header="Position ID")
    asset_type = CharColumn(header="Asset type")
    nwa = CharColumn(header="NWA")

class DividendsSheet(TableSheet):
    # 2024: Date of Payment	Instrument Name	Net Dividend Received (USD)	Net Dividend Received (EUR)	Withholding Tax Rate (%)	Withholding Tax Amount (USD)
    #       Withholding Tax Amount (EUR)	Position ID	Type	ISIN
    date = CharColumn(header="Date of Payment")
    name = CharColumn(header="Instrument Name")
    net_dividend = CharColumn(header="Net Dividend Received (USD)")
    net_dividend_eur = CharColumn(header="Net Dividend Received (EUR)")
    withholding_tax_rate = CharColumn(header="Withholding Tax Rate (%)")
    withholding_tax_amount = CharColumn(header="Withholding Tax Amount (USD)")
    withholding_tax_amount_eur = CharColumn(header="Withholding Tax Amount (EUR)")
    position_id = CharColumn(header="Position ID")
    type = CharColumn(header="Type")
    isin = CharColumn(header="ISIN")

class EToroWorkbook(TemplatedWorkbook):
    closed_positions = ClosedPositionsSheet(sheetname='Closed Positions')
    transactions = AccountActivityReportSheet(sheetname='Account Activity')
    dividends = DividendsSheet(sheetname='Dividends')

class CompanyInfoSheet(TableSheet):
    symbol = CharColumn(header='Symbol')
    ISIN = CharColumn(header='ISIN')
    name = CharColumn(header='Name')
    address = CharColumn(header='Address')
    country_code = CharColumn(header='CountryCode')

class CompanyWorkbook(TemplatedWorkbook):
    info = CompanyInfoSheet(sheetname='Info')

class DividendsOutputSheet(TableSheet):
    skipped = CharColumn(header="Skipped", width=7)
    date = CharColumn(header="Date", width=12)
    symbol = CharColumn(header="Symbol", width=12)
    ISIN = CharColumn(header="ISIN")
    name = CharColumn(header="Company/Name", width=50)
    address = CharColumn(header="Address", width=65)
    country = CharColumn(header="CountryCode", width=7)
    netto_dividend_eur = CharColumn(header="Netto dividend [EUR]")
    dividend_tax_eur = CharColumn(header="Withholding Tax Amount [EUR]")
    dividend_eur = CharColumn(header="Gross dividend [EUR]")
    #currency = CharColumn(header="Orig. currency")
    position_ids = CharColumn(header="Position ID(s)", width=100)

class DividendsOutputWorkbook(TemplatedWorkbook):
    dividends = DividendsOutputSheet()

# returns [date_format, float_with_comma]
def determine_date_format_and_comma(date):
    try:
        datetime.datetime.strptime(date, ETORO_DATETIME_FORMAT_EN1)
        return [ETORO_DATETIME_FORMAT_EN1, False]
    except ValueError:
        pass
    try:
        datetime.datetime.strptime(date, ETORO_DATETIME_FORMAT_SL1)
        return [ETORO_DATETIME_FORMAT_SL1, True]
    except ValueError:
        pass
    try:
        datetime.datetime.strptime(date, ETORO_DATETIME_FORMAT_EN2)
        return [ETORO_DATETIME_FORMAT_EN2, False]
    except ValueError:
        pass
    try:
        datetime.datetime.strptime(date, ETORO_DATETIME_FORMAT_SL2)
        return [ETORO_DATETIME_FORMAT_SL2, True]
    except ValueError:
        print("ERROR: Could not determine eToro DATETIME format!")
        sys.exit(-1)


def get_exchange_rate(rates, trade_date, currency):
    date = trade_date.strftime("%Y%m%d")
    if date in rates:
        return float(rates[date][currency])
    else:
        for i in range(0, 6):
            trade_date = trade_date - datetime.timedelta(days=1)
            date = trade_date.strftime("%Y%m%d")
            if date in rates:
                return float(rates[date][currency])
            if i == 6:
                sys.exit(
                    "Error: There is no exchange rate for " + str(date)
                )

def get_position_symbols(transactionList):
    syms = {}
    for transactionSheet in transactionList:
        if transactionSheet is None:
            continue

        for xlsTransaction in transactionSheet:
            # Date	Account Balance	Type	Details	Position ID	Amount	Realized Equity Change	Realized Equity	NWA
            #if xlsTransaction.type == "Open Position":
            if xlsTransaction.position_id is None or xlsTransaction.details is None or xlsTransaction.details.find("/") < 0:
                continue
            details_split = xlsTransaction.details.split("/", 1)
            position_id = int(xlsTransaction.position_id)
            syms[position_id] = details_split[0].upper()
    return syms

def update_position_symbols_from_dividends(dividendsList, companyList, syms):
    for diviSheet in dividendsList:
        if diviSheet is None:
            continue

        for xlsDividend in diviSheet:
            position_id = int(xlsDividend.position_id)
            if syms.get(position_id) is not None:
                continue

            companyInfo = get_company_info_by_isin(xlsDividend.isin, companyList)
            if companyInfo is None:
                print("!!! POZOR / NAPAKA: Ključa [position_id={0}] ni v slovarjih [openPositions, Company_info.xlxs]!".format(position_id))
                print("                    Izvozi account statement za daljše obdobje oz. dodaj podatke za [ISIN={0}] v Company_info.xlxs.".format(xlsDividend.isin))
                sys.exit(1)

            syms[position_id] = companyInfo.symbol
            # DEBUG print("!!! Found {0}: {1}, {2}".format(position_id, companyInfo.symbol, companyInfo.name))
    return syms

def get_company_info(symbol, companyList):
    symbol = symbol.upper()
    for companyInfo in companyList:
        if companyInfo.symbol == symbol:
            return companyInfo
    return None

def get_company_info_by_isin(isin, companyList):
    for companyInfo in companyList:
        if companyInfo.ISIN == isin:
            return companyInfo
    return None

def str2float(num):
    global float_with_comma
    if float_with_comma:
        return float(num.replace(",", "."))
    return float(num)

# noinspection PyUnusedLocal
def main():
    global float_with_comma

    print("------------------------------------------------------------------------------")
    print("| eToro->eDavki | verzija " + APP_VER)
    print("------------------------------------------------------------------------------")

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "eToroXLSXFiles",
        metavar="eToro-xlsx-file",
        help="eToro XLSX datoteka (\"XLSX Statement\")",
        nargs="+",
    )
    parser.add_argument(
        "-y",
        metavar="report-year",
        type=int,
        default=0,
        help="Datoteke bodo generirane za izbrano leto (privzeto za " + str(datetime.date.today().year - 1) + ")",
    )
    parser.add_argument(
        "-t",
        help="Testing",
        action="store_true",
    )
    parser.add_argument(
        "-c",
        help="(Doh-KDVP) Vključi tudi kripto pozicije brez vzvoda v poročilu (običajno za s.p.; d.o.o.). Kripto pozicije z vzvodom (CFD) so vedno vključene.",
        action="store_true",
        default=False
    )

    args = parser.parse_args()
    inputFilenames = args.eToroXLSXFiles
    if args.y == 0:
        reportYear = datetime.date.today().year - 1
    else:
        reportYear = int(args.y)

    reportCryptos = args.c

    test = args.t


    if not os.path.isfile("taxpayer.xml"):
        print("Doh-Div.xml potrebuje tvojo davčno številko. Če se zmotiš, jo lahko spremeniš ročno (taxpayer.xml) ali pa kar pobrišeš taxpayer.xml in ponovno poženeš program.")
        tax_number = input("Vnesi svojo davčno številko: ")
        taxpayer_type = input("Tip davkoplačevalca (običajno FO, možnosti: FO - fizična oseba, PO - pravna oseba, SP - fizična oseba z dejavnostjo): ") or "FO"
        taxpayer_type = taxpayer_type.upper()
        f = open("taxpayer.xml", "w+", encoding="utf-8")
        f.write(
            "<taxpayer>\n"
            "   <taxNumber>" + tax_number + "</taxNumber>\n"
            "   <taxpayerType>" + taxpayer_type + "</taxpayerType>\n"
            "</taxpayer>"
        )
        f.close()

    if not os.path.isdir("output"):
        os.mkdir("output")


    """ Parse taxpayer information from the local taxpayer.xml file """
    taxpayer = xml.etree.ElementTree.parse("taxpayer.xml").getroot()
    taxpayerConfig = {
        "taxNumber": taxpayer.find("taxNumber").text,
        "taxpayerType": "FO",
    }

    """ Creating daily exchange rates object """
    bsRateXmlFilename = ("bsrate-" + str(datetime.date.today().year) + str(datetime.date.today().month) + str(datetime.date.today().day) + ".xml")
    if not os.path.isfile(bsRateXmlFilename):
        for file in glob.glob("bsrate-*.xml"):
            os.remove(file)
        urllib.request.urlretrieve(bsRateXmlUrl, bsRateXmlFilename)
    bsRateXml = xml.etree.ElementTree.parse(bsRateXmlFilename).getroot()

    rates = {}
    for d in bsRateXml:
        date = d.attrib["datum"].replace("-", "")
        rates[date] = {}
        for r in d:
            currency = r.attrib["oznaka"]
            rates[date][currency] = r.text

    """ Load company info """
    companyList = list(CompanyWorkbook(file="Company_info.xlsx").info.read())

    """ Parsing of XLSX files """
    tradesList = []
    transactionList = []
    dividendsList = []
    for filename in inputFilenames:
        wb = EToroWorkbook(file=filename)
        tradesList.append(list(wb.closed_positions.read()))
        transactionList.append(list(wb.transactions.read()))
        dividendsList.append(list(wb.dividends.read()))

    statementStartDate = datetime.datetime(year=reportYear, month=1, day=1)
    statementEndDate = datetime.datetime(year=reportYear, month=12, day=31)

    """ Dictionary of stock trade arrays, each key represents a group of trades of same resource """
    longNormalTrades = {}
    shortNormalTrades = {}
    longDerivateTrades = {}
    shortDerivateTrades = {}
    skippedCryptoTrades = {}

    """ Get trades from the worksheet and sort them by PositionID """
    ETORO_DATETIME_FORMAT = None

    allTradesByPositionID = {}
    allTradesBySymbol = {}
    positionSymbols = get_position_symbols(transactionList)
    positionSymbols = update_position_symbols_from_dividends(dividendsList, companyList, positionSymbols)
    for tradeSheet in tradesList:
        if tradeSheet is None:
            continue

        for xlsTrade in tradeSheet:
            # determine etoro datetime format
            if ETORO_DATETIME_FORMAT is None:
                ETORO_DATETIME_FORMAT, float_with_comma = determine_date_format_and_comma(xlsTrade.close_date)

            close_date = datetime.datetime.strptime(xlsTrade.close_date, ETORO_DATETIME_FORMAT)
            if close_date.year != reportYear:
                # print("Skipping trade (year: " + str(close_date.year) + "): " + str(xlsTrade))
                continue

            open_date = datetime.datetime.strptime(xlsTrade.open_date, ETORO_DATETIME_FORMAT)  # ex.: 02/06/2020 13:57

            action = xlsTrade.action.split(" ", 1)
            buy_sell = action[0]
            position_id = int(xlsTrade.position_id)
            name = action[1]

            symbol = positionSymbols[position_id] if position_id in positionSymbols else None

            # fix for forex symbols
            if name is not None and len(name) == 7 and name[:4] == symbol + "/":
                symbol = name[0:3]+name[4:]

            ifi_type = xlsTrade.type

            try:
                leverage = int(xlsTrade.leverage) if xlsTrade.leverage is not None else 0
            except ValueError:
                leverage = 1

            if leverage is not None and leverage > 1:
                amount = str2float(xlsTrade.amount) * leverage
            else:
                amount = str2float(xlsTrade.amount)
            units = str2float(xlsTrade.units)
            profit = str2float(xlsTrade.profit)

            # open & close prices are bogus in eToro statement... calculate it from amount and profit
            #open_price = str2float(xlsTrade.open_rate)
            #close_price = str2float(xlsTrade.close_rate)

            open_price = amount / units
            close_price = (amount + profit) / units


            open_rate = get_exchange_rate(rates, open_date, ETORO_CURRENCY)
            if get_exchange_rate(rates, close_date, ETORO_CURRENCY) is None:
                print("")
            close_rate = get_exchange_rate(rates, close_date, ETORO_CURRENCY)

            open_price_eur = open_price / open_rate
            close_price_eur = close_price / close_rate

            if buy_sell == "Buy":
                position_type = "long"
            elif buy_sell == "Sell":
                position_type = "short"
            else:
                print("ERROR: Could not determine position type! ")
                sys.exit(-1)

            if ifi_type in derivateAssets:
                asset_type = "derivate"
            elif ifi_type in normalAssets:
                if leverage > 1:
                    print("ERROR: Leverage > 1 but asset type is not a derivate: {0}. Please report it on github.".format(ifi_type))
                    sys.exit(-1)
                asset_type = "normal"
            else:
                print("ERROR: Unknown asset type: {0}. Please report it on github.".format(ifi_type))
                sys.exit(-1)

            is_etf = ifi_type == "ETF"

            trade_open = {
                "position_id": position_id,
                "symbol": symbol,
                "position_type": position_type,
                "name": name,
                "is_etf": is_etf,
                "ifi_type": ifi_type,
                "leverage": leverage,
                "asset_type": asset_type,
                "quantity": units,
                "trade_date": open_date,
                "trade_price_eur": open_price_eur,
                "isin": xlsTrade.isin,

                # extra info
                "open_price_eur": open_price_eur,
                "close_price_eur": close_price_eur,
                "open_date": open_date,
                "close_date": close_date,
            }

            trade_close = {
                "position_id": position_id,
                "symbol": symbol,
                "position_type": position_type,
                "name": name,
                "is_etf": is_etf,
                "ifi_type": ifi_type,
                "leverage": leverage,
                "asset_type": asset_type,
                "quantity": -units,
                "trade_date": close_date,
                "trade_price_eur": close_price_eur,
                "isin": xlsTrade.isin,

                # extra info
                "open_price_eur": open_price_eur,
                "close_price_eur": close_price_eur,
                "open_date": open_date,
                "close_date": close_date,
            }

            allTradesByPositionID[position_id] = trade_open
            if symbol is not None:
                allTradesBySymbol[symbol] = trade_open


            if reportCryptos == False and ifi_type == "Crypto":
                if name in skippedCryptoTrades.keys():
                    skippedCryptoTrades[name].extend([trade_open, trade_close])
                else:
                    skippedCryptoTrades[name] = [trade_open, trade_close]
                continue


            if asset_type == "normal":
                if position_type == "long":
                    if name in longNormalTrades.keys():
                        longNormalTrades[name].extend([trade_open, trade_close])
                    else:
                        longNormalTrades[name] = [trade_open, trade_close]
                elif position_type == "short":
                    if name in shortNormalTrades.keys():
                        shortNormalTrades[name].extend([trade_open, trade_close])
                    else:
                        shortNormalTrades[name] = [trade_open, trade_close]
                else:
                    print("ERROR: Could not determine position type! ")
                    sys.exit(-1)

            else:
                if position_type == "long":
                    if name in longDerivateTrades.keys():
                        longDerivateTrades[name].extend([trade_open, trade_close])
                    else:
                        longDerivateTrades[name] = [trade_open, trade_close]
                elif position_type == "short":
                    if name in shortDerivateTrades.keys():
                        shortDerivateTrades[name].extend([trade_open, trade_close])
                    else:
                        shortDerivateTrades[name] = [trade_open, trade_close]
                else:
                    print("ERROR: Could not determine position type! ")
                    sys.exit(-1)

            """ else:
                sys.exit(
                    "Error: cannot figure out if trade is Normal or Derivate, Long or Short"
                ) """

    """ Sort trades by position ID """
    for securityID in longNormalTrades:
        longNormalTrades[securityID].sort(key=itemgetter('trade_date', 'position_id'))
    for securityID in shortNormalTrades:
        shortNormalTrades[securityID].sort(key=itemgetter('trade_date', 'position_id'))
    for securityID in longDerivateTrades:
        longDerivateTrades[securityID].sort(key=itemgetter('trade_date', 'position_id'))
    for securityID in shortDerivateTrades:
        shortDerivateTrades[securityID].sort(key=itemgetter('trade_date', 'position_id'))

    for securityID in skippedCryptoTrades:
        skippedCryptoTrades[securityID].sort(key=itemgetter('trade_date', 'position_id'))



    """ Save debug info to XLS """
    wb = Workbook()
    sh = wb.create_sheet(title="Normal (long)")
    sh.append([ "Symbol", "Name", "ISIN", "Is ETF", "Action", "Trade date", "Quantity", "Trade price (EUR)" ])
    for securityID in longNormalTrades:
        trades = longNormalTrades[securityID]
        for trade in trades:
            sh.append([
                trade["symbol"],
                trade["name"],
                trade["isin"],
                "true" if trades[0]["is_etf"] else "false",
                "Open" if trade["quantity"] > 0 else "Close",
                trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT),
                trade["quantity"] if trade["quantity"] >= 0 else -trade["quantity"],
                trade["trade_price_eur"]
            ])

    sh = wb.create_sheet(title="Derivate (long)")
    sh.append([ "Symbol", "Name", "ISIN", "Is ETF", "Action", "Trade date", "Quantity", "Trade price (EUR)" ])
    for securityID in longDerivateTrades:
        trades = longDerivateTrades[securityID]
        for trade in trades:
            sh.append([
                trade["symbol"],
                trade["name"],
                trade["isin"],
                "true" if trades[0]["is_etf"] else "false",
                "Open" if trade["quantity"] > 0 else "Close",
                trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT),
                trade["quantity"] if trade["quantity"] >= 0 else -trade["quantity"],
                trade["trade_price_eur"]
            ])


    sh = wb.create_sheet(title="Derivate (short)")
    sh.append([ "Symbol", "Name", "ISIN", "Is ETF", "Action", "Trade date", "Quantity", "Trade price (EUR)" ])
    for securityID in shortDerivateTrades:
        trades = shortDerivateTrades[securityID]
        for trade in trades:
            sh.append([
                trade["symbol"],
                trade["name"],
                trade["isin"],
                "true" if trades[0]["is_etf"] else "false",
                "Open" if trade["quantity"] > 0 else "Close",
                trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT),
                trade["quantity"] if trade["quantity"] >= 0 else -trade["quantity"],
                trade["trade_price_eur"]
            ])


    sh = wb.create_sheet(title="Skipped crypto")
    sh.append([ "Symbol", "Name", "Action", "Trade date", "Quantity", "Trade price (EUR)" ])
    for securityID in skippedCryptoTrades:
        trades = skippedCryptoTrades[securityID]
        for trade in trades:
            sh.append([
                trade["symbol"],
                trade["name"],
                "Open" if trade["quantity"] > 0 else "Close",
                trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT),
                trade["quantity"] if trade["quantity"] >= 0 else -trade["quantity"],
                trade["trade_price_eur"]
            ])

    filename = "output/Debug-{0}.xlsx".format(reportYear)
    wb.save(filename)
    print("{0} created ".format(filename))



    ###########
    ########### Doh-KDVP
    ###########

    """ Generate the files for Normal """
    envelope = xml.etree.ElementTree.Element("Envelope", xmlns="http://edavki.durs.si/Documents/Schemas/Doh_KDVP_9.xsd")
    envelope.set("xmlns:edp", "http://edavki.durs.si/Documents/Schemas/EDP-Common-1.xsd")
    header = xml.etree.ElementTree.SubElement(envelope, "edp:Header")
    taxpayer = xml.etree.ElementTree.SubElement(header, "edp:taxpayer")
    xml.etree.ElementTree.SubElement(taxpayer, "edp:taxNumber").text = taxpayerConfig["taxNumber"]
    xml.etree.ElementTree.SubElement(taxpayer, "edp:taxpayerType").text = taxpayerConfig["taxpayerType"]
    Workflow = xml.etree.ElementTree.SubElement(header, "edp:Workflow")
    if test:
        xml.etree.ElementTree.SubElement(Workflow, "edp:DocumentWorkflowID").text = "I"
    else:
        xml.etree.ElementTree.SubElement(Workflow, "edp:DocumentWorkflowID").text = "O"
    xml.etree.ElementTree.SubElement(envelope, "edp:AttachmentList")
    xml.etree.ElementTree.SubElement(envelope, "edp:Signatures")

    body = xml.etree.ElementTree.SubElement(envelope, "body")
    xml.etree.ElementTree.SubElement(body, "edp:bodyContent")
    Doh_KDVP = xml.etree.ElementTree.SubElement(body, "Doh_KDVP")
    KDVP = xml.etree.ElementTree.SubElement(Doh_KDVP, "KDVP")
    if test:
        xml.etree.ElementTree.SubElement(KDVP, "DocumentWorkflowID").text = "I"
    else:
        xml.etree.ElementTree.SubElement(KDVP, "DocumentWorkflowID").text = "O"
    xml.etree.ElementTree.SubElement(KDVP, "Year").text = str(reportYear)
    xml.etree.ElementTree.SubElement(KDVP, "PeriodStart").text = statementStartDate.strftime(EDAVKI_DATETIME_FORMAT)
    xml.etree.ElementTree.SubElement(KDVP, "PeriodEnd").text = statementEndDate.strftime(EDAVKI_DATETIME_FORMAT)
    xml.etree.ElementTree.SubElement(KDVP, "IsResident").text = "true"
    xml.etree.ElementTree.SubElement(KDVP, "SecurityCount").text = str(len(longNormalTrades))
    xml.etree.ElementTree.SubElement(KDVP, "SecurityShortCount").text = str(len(shortNormalTrades))
    xml.etree.ElementTree.SubElement(KDVP, "SecurityWithContractCount").text = "0"
    xml.etree.ElementTree.SubElement(KDVP, "SecurityWithContractShortCount").text = "0"
    xml.etree.ElementTree.SubElement(KDVP, "ShareCount").text = "0"

    for securityID in longNormalTrades:
        trades = longNormalTrades[securityID]
        KDVPItem = xml.etree.ElementTree.SubElement(Doh_KDVP, "KDVPItem")
        InventoryListType = xml.etree.ElementTree.SubElement(KDVPItem, "InventoryListType").text = "PLVP"
        Name = xml.etree.ElementTree.SubElement(KDVPItem, "Name").text = trades[0]["name"]
        HasForeignTax = xml.etree.ElementTree.SubElement(KDVPItem, "HasForeignTax").text = "false"
        HasLossTransfer = xml.etree.ElementTree.SubElement(KDVPItem, "HasLossTransfer").text = "false"
        ForeignTransfer = xml.etree.ElementTree.SubElement(KDVPItem, "ForeignTransfer").text = "false"
        TaxDecreaseConformance = xml.etree.ElementTree.SubElement(KDVPItem, "TaxDecreaseConformance").text = "false"
        Securities = xml.etree.ElementTree.SubElement(KDVPItem, "Securities")
        # We need to enter either ISIN, Code or Name
        # ISIN = xml.etree.ElementTree.SubElement(Securities, "ISIN").text = trades[0]["isin"]
        if len(trades) > 0 and "symbol" in trades[0] and trades[0]["symbol"] is not None:
            Code = xml.etree.ElementTree.SubElement(Securities, "Code").text = trades[0]["symbol"][:10]
        Name = xml.etree.ElementTree.SubElement(Securities, "Name").text = trades[0]["name"]
        IsFond = xml.etree.ElementTree.SubElement(Securities, "IsFond").text = "true" if trades[0]["is_etf"] else "false"

        F8Value = 0
        n = -1
        for trade in trades:
            n += 1
            Row = xml.etree.ElementTree.SubElement(Securities, "Row")
            ID = xml.etree.ElementTree.SubElement(Row, "ID").text = str(n)
            if trade["quantity"] > 0:
                PurchaseSale = xml.etree.ElementTree.SubElement(Row, "Purchase")
                # Datum pridobitve
                F1 = xml.etree.ElementTree.SubElement(PurchaseSale, "F1").text = trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT)
                # Način pridobitve: A - vložek kapitala, B - nakup, C - povečanje kapitala družbe z lastnimi sredstvi zavezanca,
                # D - povečanje kapitala družbe iz sredstev družbe, E - zamenjava kapitala ob statusnih spremembah družbe, F - dedovanje,
                # G - darilo, H - drugo, I - povečanje kapitalskega deleža v osebni družbi zaradi pripisa dobička kapitalskemu deležu
                F2 = xml.etree.ElementTree.SubElement(PurchaseSale, "F2").text = "B"
                # Količina
                F3 = xml.etree.ElementTree.SubElement(PurchaseSale, "F3").text = "{0:.8f}".format(trade["quantity"])
                # Nabavna vrednost ob pridobitvi (na enoto)
                F4 = xml.etree.ElementTree.SubElement(PurchaseSale, "F4").text = "{0:.8f}".format(trade["trade_price_eur"])
                # Plačan davek na dediščine in darila (F2 == F | G)
                F5 = xml.etree.ElementTree.SubElement(PurchaseSale, "F5").text = "0.0000"
            elif trade["quantity"] == 0:
                print("Error! Trade units == 0! " + str(trade))
            else:
                PurchaseSale = xml.etree.ElementTree.SubElement(Row, "Sale")
                # Datum odsvojitve
                F6 = xml.etree.ElementTree.SubElement(PurchaseSale, "F6").text = trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT)
                # Količina odsvojenega v.p.
                F7 = xml.etree.ElementTree.SubElement(PurchaseSale, "F7").text = "{0:.8f}".format(-trade["quantity"])
                # Vrednost ob osvojitvi (na enoto)
                F9 = xml.etree.ElementTree.SubElement(PurchaseSale, "F9").text = "{0:.8f}".format(trade["trade_price_eur"])
                # Pravilo iz drugega odstavka v povezavi s petim odstavkom 97.člena ZDoh-2
                # TODO:
                #F10 = xml.etree.ElementTree.SubElement(PurchaseSale, "F10").text = "NE"
            # Trenutna zaloga
            F8Value += trade["quantity"]
            F8 = xml.etree.ElementTree.SubElement(Row, "F8").text = "{0:.8f}".format(F8Value)
        # trades
    # longNormalTrades

    for securityID in shortNormalTrades:
        trades = shortNormalTrades[securityID]
        KDVPItem = xml.etree.ElementTree.SubElement(Doh_KDVP, "KDVPItem")
        InventoryListType = xml.etree.ElementTree.SubElement(KDVPItem, "InventoryListType").text = "PLVPSHORT"
        Name = xml.etree.ElementTree.SubElement(KDVPItem, "Name").text = trades[0]["name"]
        HasForeignTax = xml.etree.ElementTree.SubElement(KDVPItem, "HasForeignTax").text = "false"
        HasLossTransfer = xml.etree.ElementTree.SubElement(KDVPItem, "HasLossTransfer").text = "false"
        ForeignTransfer = xml.etree.ElementTree.SubElement(KDVPItem, "ForeignTransfer").text = "false"
        TaxDecreaseConformance = xml.etree.ElementTree.SubElement(KDVPItem, "TaxDecreaseConformance").text = "false"
        SecuritiesShort = xml.etree.ElementTree.SubElement(KDVPItem, "SecuritiesShort")
        # We need to enter either ISIN, Code or Name
        #ISIN = xml.etree.ElementTree.SubElement(SecuritiesShort, "ISIN").text = trades[0]["isin"]
        if len(trades) > 0 and "symbol" in trades[0] and trades[0]["symbol"] is not None:
            Code = xml.etree.ElementTree.SubElement(SecuritiesShort, "Code").text = trades[0]["symbol"][:10]
        Name = xml.etree.ElementTree.SubElement(SecuritiesShort, "Name").text = trades[0]["name"]
        IsFond = xml.etree.ElementTree.SubElement(SecuritiesShort, "IsFond").text = "true" if trades[0]["is_etf"] else "false"

        F8Value = 0
        n = -1
        for trade in trades:
            n += 1
            Row = xml.etree.ElementTree.SubElement(SecuritiesShort, "Row")
            ID = xml.etree.ElementTree.SubElement(Row, "ID").text = str(n)
            if trade["quantity"] > 0:
                PurchaseSale = xml.etree.ElementTree.SubElement(Row, "Purchase")
                F1 = xml.etree.ElementTree.SubElement(PurchaseSale, "F1").text = trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT)
                F2 = xml.etree.ElementTree.SubElement(PurchaseSale, "F2").text = "A"
                F3 = xml.etree.ElementTree.SubElement(PurchaseSale, "F3").text = "{0:.8f}".format(trade["quantity"])
                F4 = xml.etree.ElementTree.SubElement(PurchaseSale, "F4").text = "{0:.8f}".format(trade["trade_price_eur"])
                F5 = xml.etree.ElementTree.SubElement(PurchaseSale, "F5").text = "0.0000"
            else:
                PurchaseSale = xml.etree.ElementTree.SubElement(Row, "Sale")
                F6 = xml.etree.ElementTree.SubElement(PurchaseSale, "F6").text = trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT)
                F7 = xml.etree.ElementTree.SubElement(PurchaseSale, "F7").text = "{0:.8f}".format(-trade["quantity"])
                F9 = xml.etree.ElementTree.SubElement(PurchaseSale, "F9").text = "{0:.8f}".format(trade["trade_price_eur"])
                # Pravilo iz drugega odstavka v povezavi s petim odstavkom 97.člena ZDoh-2
                # TODO:
                # F10 = xml.etree.ElementTree.SubElement(PurchaseSale, "F10").text = "NE"
            # Trenutna zaloga
            F8Value += trade["quantity"]
            F8 = xml.etree.ElementTree.SubElement(Row, "F8").text = "{0:.8f}".format(F8Value)
        # trades
    # shortNormalTrades

    xmlString = xml.etree.ElementTree.tostring(envelope)
    prettyXmlString = minidom.parseString(xmlString).toprettyxml(indent="\t")
    with open("output/Doh-KDVP.xml", "w", encoding="utf-8") as f:
        f.write(prettyXmlString)
        print("output/Doh-KDVP.xml created")


    print("")

    for securityID in skippedCryptoTrades:
        trades = skippedCryptoTrades[securityID]
        ids = []
        name = trades[0]["name"]
        symbol = trades[0]["symbol"]
        for trade in trades:
            if trade["position_id"] not in ids:
                ids.append(trade["position_id"])

        ids = ','.join(map(str, ids))
        print("Crypto: skipped {0}/{1} ({2})".format(name, symbol, ids))

    print("")

    ###########
    ########### D-IFI
    ###########

    """ Generate the files for Derivates """
    envelope = xml.etree.ElementTree.Element("Envelope", xmlns="http://edavki.durs.si/Documents/Schemas/D_IFI_4.xsd")
    envelope.set("xmlns:edp", "http://edavki.durs.si/Documents/Schemas/EDP-Common-1.xsd")
    header = xml.etree.ElementTree.SubElement(envelope, "edp:Header")
    taxpayer = xml.etree.ElementTree.SubElement(header, "edp:taxpayer")
    xml.etree.ElementTree.SubElement(taxpayer, "edp:taxNumber").text = taxpayerConfig["taxNumber"]
    xml.etree.ElementTree.SubElement(taxpayer, "edp:taxpayerType").text = taxpayerConfig["taxpayerType"]
    Workflow = xml.etree.ElementTree.SubElement(header, "edp:Workflow")
    if test:
        xml.etree.ElementTree.SubElement(Workflow, "edp:DocumentWorkflowID").text = "I"
    else:
        xml.etree.ElementTree.SubElement(Workflow, "edp:DocumentWorkflowID").text = "O"
    xml.etree.ElementTree.SubElement(envelope, "edp:AttachmentList")
    xml.etree.ElementTree.SubElement(envelope, "edp:Signatures")


    body = xml.etree.ElementTree.SubElement(envelope, "body")
    xml.etree.ElementTree.SubElement(body, "edp:bodyContent")
    difi = xml.etree.ElementTree.SubElement(body, "D_IFI")
    xml.etree.ElementTree.SubElement(difi, "PeriodStart").text = statementStartDate.strftime(EDAVKI_DATETIME_FORMAT)
    xml.etree.ElementTree.SubElement(difi, "PeriodEnd").text = statementEndDate.strftime(EDAVKI_DATETIME_FORMAT)
    xml.etree.ElementTree.SubElement(difi, "TelephoneNumber").text = ""
    xml.etree.ElementTree.SubElement(difi, "Email").text = ""

    n = 0
    for securityID in longDerivateTrades:
        trades = longDerivateTrades[securityID]
        n += 1

        TItem = xml.etree.ElementTree.SubElement(difi, "TItem")
        TypeId = xml.etree.ElementTree.SubElement(TItem, "TypeId").text = "PLIFI"
        if trades[0]["ifi_type"] == "FUT":
            Type = xml.etree.ElementTree.SubElement(TItem, "Type").text = "01"
            TypeName = xml.etree.ElementTree.SubElement(TItem, "TypeName").text = "terminska pogodba"
        elif trades[0]["ifi_type"] == "CFD":
            Type = xml.etree.ElementTree.SubElement(TItem, "Type").text = "02"
            TypeName = xml.etree.ElementTree.SubElement(TItem, "TypeName").text = "finančne pogodbe na razliko"
        elif trades[0]["ifi_type"] == "OPT":
            Type = xml.etree.ElementTree.SubElement(TItem, "Type").text = "03"
            TypeName = xml.etree.ElementTree.SubElement(TItem, "TypeName").text = "opcija in certifikat"
        else:
            Type = xml.etree.ElementTree.SubElement(TItem, "Type").text = "04"
            TypeName = xml.etree.ElementTree.SubElement(TItem, "TypeName").text = "drugo"

        Name = xml.etree.ElementTree.SubElement(TItem, "Name").text = trades[0]["name"]
        if len(trades) > 0 and "symbol" in trades[0] and trades[0]["symbol"] is not None:
            Code = xml.etree.ElementTree.SubElement(TItem, "Code").text = trades[0]["symbol"]
        #ISIN = xml.etree.ElementTree.SubElement(TItem, "ISIN").text = trades[0]["isin"]
        HasForeignTax = xml.etree.ElementTree.SubElement(TItem, "HasForeignTax").text = "false"

        F8Value = 0
        for trade in trades:
            TSubItem = xml.etree.ElementTree.SubElement(TItem, "TSubItem")
            if trade["quantity"] > 0:
                PurchaseSale = xml.etree.ElementTree.SubElement(TSubItem, "Purchase")
                # Datum pridobitve
                F1 = xml.etree.ElementTree.SubElement(PurchaseSale, "F1").text = trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT)
                # Način pridobitve: A - nakup, B - dedovanje, C - darila, D - drugo
                F2 = xml.etree.ElementTree.SubElement(PurchaseSale, "F2").text = "A"
                # Količina
                F3 = xml.etree.ElementTree.SubElement(PurchaseSale, "F3").text = "{0:.8f}".format(trade["quantity"])
                # Nabavna vrednost ob pridobitvi (na enoto)
                F4 = xml.etree.ElementTree.SubElement(PurchaseSale, "F4").text = "{0:.8f}".format(trade["trade_price_eur"])
                # Trgovanje z vzvodom
                F9 = xml.etree.ElementTree.SubElement(PurchaseSale, "F9").text = "true" if trade["leverage"] > 1 else "false"
            else:
                PurchaseSale = xml.etree.ElementTree.SubElement(TSubItem, "Sale")
                # Datum odsvojitve
                F5 = xml.etree.ElementTree.SubElement(PurchaseSale, "F5").text = trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT)
                # Količina odsvojenega v.p.
                F6 = xml.etree.ElementTree.SubElement(PurchaseSale, "F6").text = "{0:.8f}".format(-trade["quantity"])
                # Vrednost ob odsvojitvi
                F7 = xml.etree.ElementTree.SubElement(PurchaseSale, "F7").text = "{0:.8f}".format(trade["trade_price_eur"])
            F8Value += trade["quantity"]
            F8 = xml.etree.ElementTree.SubElement(TSubItem, "F8").text = "{0:.8f}".format(F8Value)
        # trades
    # longDerivateTrades

    for securityID in shortDerivateTrades:
        trades = shortDerivateTrades[securityID]
        n += 1

        TItem = xml.etree.ElementTree.SubElement(difi, "TItem")
        TypeId = xml.etree.ElementTree.SubElement(TItem, "TypeId").text = "PLIFIShort"
        if trades[0]["ifi_type"] == "FUT":
            Type = xml.etree.ElementTree.SubElement(TItem, "Type").text = "01"
            TypeName = xml.etree.ElementTree.SubElement(TItem, "TypeName").text = "terminska pogodba"
        elif trades[0]["ifi_type"] == "CFD":
            Type = xml.etree.ElementTree.SubElement(TItem, "Type").text = "02"
            TypeName = xml.etree.ElementTree.SubElement(TItem, "TypeName").text = "finančne pogodbe na razliko"
        elif trades[0]["ifi_type"] == "OPT":
            Type = xml.etree.ElementTree.SubElement(TItem, "Type").text = "03"
            TypeName = xml.etree.ElementTree.SubElement(TItem, "TypeName").text = "opcija in certifikat"
        else:
            Type = xml.etree.ElementTree.SubElement(TItem, "Type").text = "04"
            TypeName = xml.etree.ElementTree.SubElement(TItem, "TypeName").text = "drugo"
        Name = xml.etree.ElementTree.SubElement(TItem, "Name").text = trades[0]["name"]
        if len(trades) > 0 and "symbol" in trades[0] and trades[0]["symbol"] is not None:
            Code = xml.etree.ElementTree.SubElement(TItem, "Code").text = trades[0]["symbol"]
        #ISIN = xml.etree.ElementTree.SubElement(TItem, "ISIN").text = trades[0]["isin"]
        HasForeignTax = xml.etree.ElementTree.SubElement(TItem, "HasForeignTax").text = "false"

        F8Value = 0
        for trade in trades:
            TShortSubItem = xml.etree.ElementTree.SubElement(TItem, "TShortSubItem")
            if trade["quantity"] > 0:
                PurchaseSale = xml.etree.ElementTree.SubElement(TShortSubItem, "Sale")
                F1 = xml.etree.ElementTree.SubElement(PurchaseSale, "F1").text = trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT)
                F2 = xml.etree.ElementTree.SubElement(PurchaseSale, "F2").text = "{0:.8f}".format(trade["quantity"])
                F3 = xml.etree.ElementTree.SubElement(PurchaseSale, "F3").text = "{0:.8f}".format(trade["trade_price_eur"])
                F9 = xml.etree.ElementTree.SubElement(PurchaseSale, "F9").text = "true" if trade["leverage"] > 1 else "false"
            else:
                PurchaseSale = xml.etree.ElementTree.SubElement(TShortSubItem, "Purchase")
                F4 = xml.etree.ElementTree.SubElement(PurchaseSale, "F4").text = trade["trade_date"].strftime(EDAVKI_DATETIME_FORMAT)
                F5 = xml.etree.ElementTree.SubElement(PurchaseSale, "F5").text = "A"
                F6 = xml.etree.ElementTree.SubElement(PurchaseSale, "F6").text = "{0:.8f}".format(-trade["quantity"])
                F7 = xml.etree.ElementTree.SubElement(PurchaseSale, "F7").text = "{0:.8f}".format(trade["trade_price_eur"])
            F8Value += trade["quantity"]
            F8 = xml.etree.ElementTree.SubElement(TShortSubItem, "F8").text = "{0:.8f}".format(F8Value)
        # trades
    # shortDerivateTrades

    xmlString = xml.etree.ElementTree.tostring(envelope)
    prettyXmlString = minidom.parseString(xmlString).toprettyxml(indent="\t")
    with open("output/D-IFI.xml", "w", encoding="utf-8") as f:
        f.write(prettyXmlString)
        print("output/D-IFI.xml created")

    ###########
    ########### Doh-Div
    ###########

    """ Get dividends from XLSX """
    dividends = []
    ETORO_DATETIME_FORMAT = None

    for diviSheet in dividendsList:
        if diviSheet is None:
            continue

        for xlsDividend in diviSheet:
            if ETORO_DATETIME_FORMAT is None:
                ETORO_DATETIME_FORMAT, float_with_comma = determine_date_format_and_comma(xlsDividend.date)

            # Date of Payment	Instrument Name	Net Dividend Received (USD)	Withholding Tax Rate (%)	Withholding Tax Amount (USD)	Position ID	Type	ISIN
            date = datetime.datetime.strptime(xlsDividend.date, ETORO_DATETIME_FORMAT)
            if date.year != reportYear:
                # print("Skipping dividend (year: " + str(date.year) + "): " + str(xlsDividend))
                continue

            position_id = int(xlsDividend.position_id)
            symbol = positionSymbols.get(position_id)

            rate = get_exchange_rate(rates, date, ETORO_CURRENCY)
            withholding_tax_rate = float(xlsDividend.withholding_tax_rate.rstrip('%')) / 100.0
            # old way
            # withholding_tax_amount = str2float(xlsDividend.withholding_tax_amount) / rate
            # gross_amount_eur = str2float(xlsDividend.net_dividend) / rate / (1 - withholding_tax_rate)
            # new (but doesn't use BSRATE)
            netto_amount_eur = str2float(xlsDividend.net_dividend_eur)
            withholding_tax_amount = str2float(xlsDividend.withholding_tax_amount_eur)
            gross_amount_eur = netto_amount_eur + withholding_tax_amount

            if symbol is None:
                print("!!! POZOR / NAPAKA: Ključa [position_id={0}] ni v slovarju [positionSymbols]!".format(position_id))
                print("                    Verjetno vhodna datoteka ne zajema celotnega obdobja obdelanih finančnih instrumentov.")
                sys.exit(1)

            dividend = {
                "position_id": position_id,
                "gross_amount_eur": gross_amount_eur,
                "netto_amount_eur": netto_amount_eur,
                "withholding_tax_amount": withholding_tax_amount,
                "withholding_tax_rate": withholding_tax_rate,
                "date": date,
                "name": xlsDividend.name,
                "symbol": symbol,
                "currency": "USD",
                "ISIN": xlsDividend.isin
            }

            dividends.append(dividend)

    ## Old version
    # openPositions = {}
    # for transactionSheet in transactionList:
    #     if transactionSheet is None:
    #         continue
    #
    #     for xlsTransaction in transactionSheet:
    #         # Date	Account Balance	Type	Details	Position ID	Amount	Realized Equity Change	Realized Equity	NWA
    #         if xlsTransaction.details is None:
    #             continue
    #
    #         if xlsTransaction.type == "Open Position" or xlsTransaction.type == "Profit/Loss of Trade":
    #             details_split = xlsTransaction.details.split("/", 1)
    #             position_id = int(xlsTransaction.position_id)
    #
    #             open_pos = {
    #                 "date": datetime.datetime.strptime(xlsTransaction.date, ETORO_DATETIME_FORMAT),
    #                 "symbol": details_split[0].upper(),
    #                 "currency": details_split[1]
    #             }
    #             openPositions[position_id] = open_pos
    #             continue
    #
    #         if xlsTransaction.details.casefold().find("dividend") < 0:
    #             continue
    #
    #         date = datetime.datetime.strptime(xlsTransaction.date, ETORO_DATETIME_FORMAT)
    #         if date.year != reportYear:
    #             # print("Skipping dividend (year: " + str(date.year) + "): " + str(xlsTransaction))
    #             continue
    #
    #         position_id = int(xlsTransaction.position_id)
    #
    #         rate = get_exchange_rate(rates, date, ETORO_CURRENCY)
    #         amount_eur = str2float(xlsTransaction.amount) / rate
    #
    #         if openPositions.get(position_id) is None:
    #             print("!!! POZOR / NAPAKA: Ključa [position_id={0}] ni v slovarju [openPositions]!".format(position_id))
    #             print("                    Verjetno vhodna datoteka ne zajema celotnega obdobja obdelanih finančnih instrumentov.")
    #             sys.exit(1)
    #
    #         open_pos = openPositions[position_id]
    #         symbol = open_pos["symbol"]
    #
    #         name = None
    #         if position_id in allTradesByPositionID:
    #             info = allTradesByPositionID[position_id]
    #             name = info["name"]
    #         if name is None and symbol in allTradesBySymbol:
    #             info = allTradesBySymbol[symbol]
    #             name = info["name"]
    #
    #
    #         dividend = {
    #             "position_id": position_id,
    #             "amount_eur": amount_eur,
    #             "date": date,
    #             "name": name,
    #             "symbol": symbol,
    #             "currency": open_pos["currency"],
    #         }
    #         dividends.append(dividend)

    """ Merge multiple dividends or payments in lieu of dividends on the same day from the same company into a single entry """
    mergedDividends = []
    for dividend in dividends:
        merged = False
        for mergedDividend in mergedDividends:
            if \
                dividend["date"].date() == mergedDividend["date"].date() \
                and dividend["symbol"] == mergedDividend["symbol"] \
                and mergedDividend["gross_amount_eur"]>=0 \
                and dividend["gross_amount_eur"]>=0 \
            :
                mergedDividend["netto_amount_eur"] = mergedDividend["netto_amount_eur"] + dividend["netto_amount_eur"]
                mergedDividend["gross_amount_eur"] = mergedDividend["gross_amount_eur"] + dividend["gross_amount_eur"]
                mergedDividend["withholding_tax_amount"] = mergedDividend["withholding_tax_amount"] + dividend["withholding_tax_amount"]
                if "positions" in mergedDividend:
                    mergedDividend["positions"].append(dividend["position_id"])
                else:
                    mergedDividend["positions"] = [mergedDividend["position_id"], dividend["position_id"]]
                merged = True
                break
        if not merged:
            mergedDividends.append(dividend)
    dividends = mergedDividends

    """ Add missing data """
    errors = []
    missing_info = []
    for dividend in dividends:
        companyInfo = get_company_info(dividend["symbol"], companyList)

        if companyInfo is not None:
            if dividend["ISIN"] != companyInfo.ISIN:
                errors.append([dividend["ISIN"], str(dividend), str(companyInfo)])

            dividend["address"] = companyInfo.address
            dividend["country"] = companyInfo.country_code
        elif not any(x["symbol"] == dividend["symbol"] for x in missing_info):
            missing_info.append({
                "symbol": dividend["symbol"],
                "ISIN": dividend["ISIN"],
                "name": dividend["name"]
            })

    if errors:
        print("!!! POZOR / NAPAKA:\n")
        for e in errors:
            print("\tISIN {0}:\n\t\t{1}\n\tse ne ujema z:\n\t\t{2}".format(e[0], e[1], e[2]))
        print("\tPreveri/popravi podatke v Company_info.xlsx in ponovno poženi program.")
        sys.exit(1)

    """ Generate Doh-Div.xml """
    envelope = xml.etree.ElementTree.Element("Envelope", xmlns="http://edavki.durs.si/Documents/Schemas/Doh_Div_3.xsd")
    envelope.set("xmlns:edp", "http://edavki.durs.si/Documents/Schemas/EDP-Common-1.xsd")
    header = xml.etree.ElementTree.SubElement(envelope, "edp:Header")
    taxpayer = xml.etree.ElementTree.SubElement(header, "edp:taxpayer")
    xml.etree.ElementTree.SubElement(taxpayer, "edp:taxNumber").text = taxpayerConfig["taxNumber"]
    xml.etree.ElementTree.SubElement(taxpayer, "edp:taxpayerType").text = taxpayerConfig["taxpayerType"]
    Workflow = xml.etree.ElementTree.SubElement(header, "edp:Workflow")
    if test:
        xml.etree.ElementTree.SubElement(Workflow, "edp:DocumentWorkflowID").text = "I"
    else:
        xml.etree.ElementTree.SubElement(Workflow, "edp:DocumentWorkflowID").text = "O"
    xml.etree.ElementTree.SubElement(envelope, "edp:AttachmentList")
    xml.etree.ElementTree.SubElement(envelope, "edp:Signatures")

    body = xml.etree.ElementTree.SubElement(envelope, "body")
    Doh_Div = xml.etree.ElementTree.SubElement(body, "Doh_Div")
    xml.etree.ElementTree.SubElement(Doh_Div, "Period").text = str(reportYear)

    for dividend in dividends:
        if round(dividend["gross_amount_eur"], 2) <= 0:
            dividend["skipped"] = "YES"
            continue

        Dividend = xml.etree.ElementTree.SubElement(body, "Dividend")
        xml.etree.ElementTree.SubElement(Dividend, "Date").text = dividend["date"].strftime(EDAVKI_DATETIME_FORMAT)

        if "ISIN" in dividend:
            xml.etree.ElementTree.SubElement(Dividend, "PayerIdentificationNumber").text = dividend["ISIN"]
        if "name" in dividend and dividend["name"] != "":
            xml.etree.ElementTree.SubElement(Dividend, "PayerName").text = dividend["name"]
        else:
            xml.etree.ElementTree.SubElement(Dividend, "PayerName").text = dividend["symbol"]
        if "address" in dividend:
            xml.etree.ElementTree.SubElement(Dividend, "PayerAddress").text = dividend["address"]
        if "country" in dividend:
            xml.etree.ElementTree.SubElement(Dividend, "PayerCountry").text = dividend["country"]
        xml.etree.ElementTree.SubElement(Dividend, "Type").text = "1"
        xml.etree.ElementTree.SubElement(Dividend, "Value").text = "{0:.2f}".format(dividend["gross_amount_eur"])
        xml.etree.ElementTree.SubElement(Dividend, "ForeignTax").text = "{0:.2f}".format(dividend["withholding_tax_amount"])
        if "country" in dividend:
            xml.etree.ElementTree.SubElement(Dividend, "SourceCountry").text = dividend["country"]
        # TODO: sestavi seznam oprostitvenih besedil (MP, clen...) iz https://www.gov.si/drzavni-organi/ministrstva/ministrstvo-za-finance/o-ministrstvu/direktorat-za-sistem-davcnih-carinskih-in-drugih-javnih-prihodkov/seznam-veljavnih-konvencij-o-izogibanju-dvojnega-obdavcevanja-dohodka-in-premozenja/
        #if "reliefStatement" in dividend:
        #    xml.etree.ElementTree.SubElement(Dividend, "ReliefStatement").text = dividend["reliefStatement"]
        #else:
        xml.etree.ElementTree.SubElement(Dividend, "ReliefStatement").text = ""

    xmlString = xml.etree.ElementTree.tostring(envelope)
    prettyXmlString = minidom.parseString(xmlString).toprettyxml(indent="\t")
    with open("output/Doh-Div.xml", "w", encoding="utf-8") as f:
        f.write(prettyXmlString)
        print("output/Doh-Div.xml created")



    ###################
    ###################
    ###################
    """ Debug output """
    rows = []
    for dividend in dividends:
        row = [
            (dividend["skipped"] if "skipped" in dividend else ""),
            dividend["date"].strftime(EDAVKI_DATETIME_FORMAT),
            dividend["symbol"],
            (dividend["ISIN"] if "ISIN" in dividend else ""),
            (dividend["name"] if not dividend["name"] is None else ""),
            (dividend["address"] if "address" in dividend else ""),
            (dividend["country"] if "country" in dividend else ""),
            "{0:.8f}".format(dividend["netto_amount_eur"]),
            "{0:.8f}".format(dividend["withholding_tax_amount"]),
            "{0:.8f}".format(dividend["gross_amount_eur"]),
            #dividend["currency"],
            dividend["position_id"] if not "positions" in dividend else ", ".join(map(str, dividend["positions"]))
        ]
        rows.append(row)


    """ Save dividend info to XLS """
    wb = DividendsOutputWorkbook(template_styles=DefaultStyleSet(
        NamedStyle(name="hyperlink")
    ))
    if len(rows) > 0:
        wb.dividends.write(
            objects=rows
        )

    filename = "output/Dividende-info-{0}.xlsx".format(reportYear)
    wb.save(filename)
    print("{0} created ".format(filename))

    print("\n------------------------------------------------------------------------------------------------------------------------------------")

    if missing_info:
        print("\n\n")
        print("------------------------------------------------------------------------------------------------------------------------------------")
        print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        print("Manjkajo podatki o podjetjih za sledeče delnice:\n")
        for mi in missing_info:
            print("{0}\t{1}\t{2}".format(mi["symbol"], mi["ISIN"], mi["name"]))

        print("Dodaj (in dopolni) zgornje podatke v Company_info.xlsx in ponovno poženi konverzijo! (Lahko nadaljuješ z oddajo, vendar bo potrebno te podatke ročno vnesti na eDavki.)")
        print("Podatke o naslovu je običajno mogoče poiskati z ISIN kodo ali simbolom na https://www.marketscreener.com/ pod zavihkom \"Company\".")
        print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        print("------------------------------------------------------------------------------------------------------------------------------------")

    sys.exit(0)


if __name__ == "__main__":
    main()
